
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 1. Criar um Excel fictício para testar
#dados = {
 #   'Colaborador': ['Tiago', 'Ana', 'Pedro', 'Marta'],
   # 'Horas': [40, 35, 42, 38]
#}
#df_inicial = pd.DataFrame(dados)
#df_inicial.to_excel('dados_equipa.xlsx', index=False)
#print("✅ Ficheiro 'dados_equipa.xlsx' criado com sucesso!")

nome= "livro1.xlsx"
try:
    def processar_pmo(nome):
# 2. Ler o ficheiro e filtrar quem trabalhou < 40h
        df = pd.read_excel(nome)
        print(f"lemos o ficheiro : {nome} ")
# abaixo_40 = df[df['Horas'] < 40]
        em_falt= df[df['dif']<0]
# print("\n--- Alerta de Horas (Abaixo de 40h) ---")
        print("Colunas encontradas no ficheiro:", df.columns.tolist())
        print(em_falt[['id ','nome ','dif']])
    # Guardar os resultados num novo ficheiro Excel
        em_falt.to_excel('urgente_rever.xlsx', index=False)
        print("\n🚀 Relatório 'urgente_rever.xlsx' gerado com sucesso!")
except FileNotFoundError:
    print("O ficheiro não foi encontrado!")
except Exception as e:
    print(f"⚠️ Ocorreu um erro inesperado: {e}")    
def corerro(nome2):
        cor_erro = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        wb = load_workbook(nome)
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            celula = ws[f'C{row}'] # Acede à célula (ex: C2, C3...)
            if celula.value<0 and celula.value is not None:
                celula.fill=cor_erro
        wb.save(nome)


try:
    nome = "livro1.xlsx"
    nome2 = "urgente_rever.xlsx"
    df= pd.read_excel(nome)
    em_falt = df[df['dif'] < 0]
    em_falt.to_excel(nome2, index=False)
    print(f"✅ Relatório '{nome2}' criado!")
    # 4. Aplicar a formatação visual
    corerro(nome2)
    print("🎨 Erros destacados a vermelho com sucesso!")

except FileNotFoundError:
    print(f"❌ Erro: Garanta que o ficheiro está na pasta.")
except Exception as e:
    print(f"⚠️ Algo correu mal: {e}")

print("\n🚀 Automação concluída! Bom almoço!")


