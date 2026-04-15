"""PMO Report Configuration Module

Centralized configuration file for the PMO (Project Management Office) reporting system.

This module defines:
- Color schemes (HEX codes) for Excel cell formatting based on task status
- Email configuration for alerts
- File paths for input/output data
- openpyxl style objects (fonts, fills) for consistent formatting

All sensitive configuration (email credentials) is loaded from environment variables.
"""

import os
from pathlib import Path
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Colors (HEX Codes) - Used for Excel status formatting
verde: str = "C6EFCE"           # Green for completed tasks
vermelho: str = "FFC7CE"        # Red for delayed tasks
amarelo: str = "FFEB9C"         # Yellow for in-progress tasks
azul_escuro: str = "000080"     # Dark blue for headers

# Color fills
from openpyxl.styles import Font, PatternFill
VERDE_FILL: PatternFill = PatternFill(start_color=verde, end_color=verde, fill_type="solid")
VERMELHO_FILL: PatternFill = PatternFill(start_color=vermelho, end_color=vermelho, fill_type="solid")
AMARELO_FILL: PatternFill = PatternFill(start_color=amarelo, end_color=amarelo, fill_type="solid")
AZUL_ESCURO_FILL: PatternFill = PatternFill(start_color=azul_escuro, end_color=azul_escuro, fill_type="solid")

# Fonts
FONTE_BRANCA: Font = Font(color="FFFFFF", bold=True)

# File paths
ARQUIVO_EXCEL_ENTRADA: str = "Report_Quinta.xlsx"
ARQUIVO_MEMORIA: Path = Path("alertas_enviados.json")
ARQUIVO_CSV: str = "dados_pmo_segunda.csv"
ARQUIVO_EXCEL_FORMATADO: str = 'Relatorio_Formatado_PMO.xlsx'

# Email settings
EMAIL_USER: str = os.getenv("EMAIL_ADDRESS", "")
EMAIL_PASSWORD: str = os.getenv("EMAIL_PASSWORD", "")

# Risk color target (red without FF prefix)
COR_ALVO: str = vermelho