"""Excel Formatting and Styling Utilities

Provides functions for applying professional formatting to Excel files including:
- Conditional color formatting based on task status
- Header styling with custom fonts and colors
- Color-coded reporting for PMO task status visualization

Uses openpyxl for low-level Excel manipulation and pandas for data integration.
"""

import logging
from pandas import DataFrame
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from config import (
    VERDE_FILL, VERMELHO_FILL, AMARELO_FILL, AZUL_ESCURO_FILL, FONTE_BRANCA,
    ARQUIVO_EXCEL_FORMATADO
)

def aplicar_cores_status(arquivo_excel_entrada: str, df: DataFrame) -> None:
    """
    Apply conditional formatting to Excel file based on task status column.

    Loads an Excel file and applies color coding to the status column based on
    DataFrame values. Also formats header row with dark blue background and white text.
    
    Color Mapping:
        - VERMELHO_FILL (Red): "Atrasado" (Delayed tasks)
        - VERDE_FILL (Green): "Concluído" (Completed tasks)
        - AMARELO_FILL (Yellow): All other statuses (In-progress, Pending, etc.)
    
    Header Styling:
        - Background: Dark blue (AZUL_ESCURO_FILL)
        - Font: White, bold (FONTE_BRANCA)

    Args:
        arquivo_excel_entrada: Path to the input Excel file to format.
                              Must be a valid .xlsx file.
        df: DataFrame containing at minimum a 'Status' column matching the Excel
            data. Row count should match Excel data rows.

    Returns:
        None: Writes formatted file to disk with side effect. Output file path
              is determined by ARQUIVO_EXCEL_FORMATADO constant.

    Raises:
        FileNotFoundError: If the input Excel file does not exist.
        openpyxl.utils.exceptions.InvalidFileException: If Excel file is corrupted.
        PermissionError: If there's no write permission for the output file.
        KeyError: If 'Status' column not found in DataFrame.

    Note:
        - Assumes Excel data starts from row 2 (row 1 is header)
        - Status column assumed to be column 3 (C)
        - Output file overwrites ARQUIVO_EXCEL_FORMATADO if it exists
        - Uses DataFrame index for row mapping, ensure index is correct

    Example:
        >>> import pandas as pd
        >>> df = pd.DataFrame({'Status': ['Atrasado', 'Concluído', 'Em Progresso']})
        >>> aplicar_cores_status('Report_Quinta.xlsx', df)
        # Applies colors to Report_Quinta.xlsx and saves as Relatorio_Formatado_PMO.xlsx
    """
    logging.info(f"Aplicar cores com base no status: {arquivo_excel_entrada}")
    wb: Workbook = load_workbook(arquivo_excel_entrada)
    ws: Worksheet = wb.active

    # Apply red for "Atrasado"
    df_riscos = df[df["Status"] == "Atrasado"]
    for index in df_riscos.index:
        ws.cell(row=index + 2, column=3).fill = VERMELHO_FILL

    # Apply green for "Concluído"
    df_concluidos = df[df["Status"] == "Concluído"]
    for index in df_concluidos.index:
        ws.cell(row=index + 2, column=3).fill = VERDE_FILL

    # Apply yellow for others
    df_outros = df[~df["Status"].isin(["Atrasado", "Concluído"])]
    for index in df_outros.index:
        ws.cell(row=index + 2, column=3).fill = AMARELO_FILL

    # Format header
    logging.info(f"A formatar visualmente o cabeçalho do relatório: {arquivo_excel_entrada}, aplicando cor azul escuro e fonte branca.")
    for i, cell in enumerate(ws[1]):
        logging.debug(f"A formatar célula do cabeçalho: {i} of {len(ws[1])} celulas.")
        cell.fill = AZUL_ESCURO_FILL
        cell.font = FONTE_BRANCA

    logging.info(f"Cores aplicadas com base no status. Relatório atualizado: {arquivo_excel_entrada}")
    logging.info(f"✨ Report Profissional gerado: {ARQUIVO_EXCEL_FORMATADO}")
    wb.save(ARQUIVO_EXCEL_FORMATADO)
