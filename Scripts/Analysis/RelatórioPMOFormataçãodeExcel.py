# Relatório PMO - Formatação de Excel
# Learning Python with a fun project: Formatting an Excel report for PMO!
# This script reads a CSV file, converts it to Excel, and applies some styling to make it look professional.
# The goal is to automate the generation of a visually appealing report that can be easily shared with stakeholders.
# Included colors based on status, bold headers, and a clean layout to enhance readability.
# Sent email with the formatted report as an attachment to ensure that the PMO team receives the latest insights without any manual effort.
# The email includes only the new risks detected in the report, making it easier for the team to focus on what matters most.

import json
import os
import sys
import smtplib
import logging
import openpyxl
import pandas as pd
import datetime
from pathlib import Path
from openpyxl import load_workbook
from email.message import EmailMessage
from openpyxl.styles import Font, PatternFill
from dotenv import load_dotenv
from typing import Optional, Dict, List, Union

from notificaçao import enviar_alerta_slack, test_api_configuration
from config import (
    ARQUIVO_EXCEL_ENTRADA, ARQUIVO_MEMORIA, ARQUIVO_CSV, ARQUIVO_EXCEL_FORMATADO,
    EMAIL_USER, EMAIL_PASSWORD, COR_ALVO
)
from data_utils import carregar_memoria, salvar_memoria, carregar_e_validar_dados, normalizar_status
from excel_formatter import aplicar_cores_status
from notifications import enviar_email
start_time = datetime.datetime.now()

# Logging configuration
# Forçar o reload da configuração para garantir que o ficheiro é escrito
for handler in logging.root.handlers[:]:
    logging.root.removeHandler(handler)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("pmo_report.log", mode='a', encoding='utf-8'), # Grava no ficheiro (Append)
        logging.StreamHandler(sys.stdout)                               # Mantém a visualização no terminal
    ]
)

def verificar_ambiente()-> bool:
    """
    Verify that critical environment files and variables are available.

    This function checks if the .env file exists, which is required for email configuration.
    It's called before starting the main processing to ensure all dependencies are met.

    Returns:
        bool: True if the environment is properly configured, False otherwise.

    Note:
        Logs an error message if the .env file is not found.
    """
    env_file = Path('.env')
    if not env_file.exists():
        logging.error("❌ Ficheiro .env não encontrado!")
        return False
    return True

logging.info("--- 🚀 Nova sessão iniciada e registada no ficheiro ---")

# Check email credentials
if EMAIL_USER is None or EMAIL_PASSWORD is None:
    logging.error("ERRO CRÍTICO: O Python não encontrou o arquivo .env ou as variáveis!")
    logging.debug(f"DEBUG: EMAIL_ADDRESS encontrado? {'Sim' if EMAIL_USER else 'Não'}")
    logging.debug(f"DEBUG: EMAIL_PASSWORD encontrado? {'Sim' if EMAIL_PASSWORD else 'Não'}")
    # Encerra o script para não dar erro de 'NoneType'
    exit()
# --- LOGICA PRINCIPAL DO AGENTE ---
def rodar_agente() -> None:
    """
    Run the PMO risk detection agent.

    This function checks the formatted Excel report for new risks (tasks marked with red color),
    compares them against previously notified risks stored in memory, and sends notifications
    for any new risks detected. It also updates the memory file with newly notified risks.

    The agent:
    1. Verifies environment configuration
    2. Loads existing notification memory
    3. Scans the Excel file for red-colored tasks (risks)
    4. Identifies new risks not previously notified
    5. Sends email and Slack notifications for new risks
    6. Updates memory with newly notified risks

    Note:
        - Assumes the Excel file has been previously formatted with color coding
        - Tasks are identified by name in column 1, status color in column 3
        - Red color (COR_ALVO) indicates high-risk tasks
        - Exits the program if environment is not properly configured

    Side effects:
        - Updates the memory JSON file
        - Sends email notifications
        - Sends Slack notifications via imported function
        - Logs detailed information about the process
    """
    if not verificar_ambiente():
        logging.error("Ambiente de execução não está configurado corretamente. O processo será encerrado.")
        exit()  # Encerra o script se o ambiente não estiver configurado, pois os próximos passos dependem de arquivos e variáveis estarem disponíveis
    memoria: Dict[str, str] = carregar_memoria()
    wb: openpyxl.Workbook = openpyxl.load_workbook(ARQUIVO_EXCEL_FORMATADO, data_only=True)
    ws: openpyxl.worksheet.worksheet.Worksheet = wb.active
    novos_alertas: List[Dict[str, str]] = []

    for row in range(2, ws.max_row + 1):
        nome: Optional[str] = ws.cell(row=row, column=1).value
        cor: str = ws.cell(row=row, column=3).fill.start_color.rgb[2:] # ter o RGB sem o prefixo 'FF'
        logging.info(f"Verificando tarefa: {nome} - Cor: {cor}")
        if cor == COR_ALVO and nome not in memoria:
            logging.info(f"Novo risco detectado: {nome}")
            novos_alertas.append({"tarefa": nome})
            memoria[nome] = "NOTIFICADO"

    if novos_alertas:
        salvar_memoria(memoria)
        logging.warning(f"{len(novos_alertas)} novos riscos encontrados. A guardar... ")
        enviar_email(novos_alertas)
        #nviar_alerta_slack(f"O script detetou {len(novos_alertas)} tarefas com status de risco no relatório de hoje. Verifica o email para detalhes.")
    else:
        # Feedback caso o dia esteja tranquilo
        logging.info("Tudo em dia! Nenhum novo risco detectado.")
        logging.info("Nenhum novo risco encontrado para notificação.")
    logging.info(f"Agente de Riscos PMO executado com sucesso!, foram detectados: {len(novos_alertas)} novos riscos.")
    memoria: Dict[str, str] = carregar_memoria()
    wb: openpyxl.Workbook = openpyxl.load_workbook(ARQUIVO_EXCEL_FORMATADO, data_only=True)
    ws: openpyxl.worksheet.worksheet.Worksheet = wb.active
    novos_alertas: List[Dict[str, str]] = []

    for row in range(2, ws.max_row + 1):
        nome: Optional[str] = ws.cell(row=row, column=1).value
        cor: str = ws.cell(row=row, column=3).fill.start_color.rgb[2:] # ter o RGB sem o prefixo 'FF'
        logging.info(f"Verificando tarefa: {nome} - Cor: {cor}")
        if cor == COR_ALVO and nome not in memoria:
            logging.info(f"Novo risco detectado: {nome}")
            novos_alertas.append({"tarefa": nome})
            memoria[nome] = "NOTIFICADO"

    if novos_alertas:
        salvar_memoria(memoria)
        logging.warning(f"{len(novos_alertas)} novos riscos encontrados. A guardar... ")
        enviar_email(novos_alertas)
        enviar_alerta_slack(f"O script detetou {len(novos_alertas)} tarefas com status de risco no relatório de hoje. Verifica o email para detalhes.")
    else:
        # Feedback caso o dia esteja tranquilo
        logging.info("Tudo em dia! Nenhum novo risco detectado.")
        logging.info("Nenhum novo risco encontrado para notificação.")
    logging.info(f"Agente de Riscos PMO executado com sucesso!, foram detectados: {len(novos_alertas)} novos riscos.")
if __name__ == "__main__":
    # Aqui colocamos a ordem dos passos que o script deve seguir
    logging.info(" Iniciando processo diário do PMO...")
    try:

        df: Optional[pd.DataFrame] = carregar_e_validar_dados(ARQUIVO_CSV)
        if df is None:
            logging.error("Falha ao carregar ou validar os dados. O processo será encerrado.")
            exit()  # Encerra o script se os dados não puderem ser carregados ou validados, pois os próximos passos dependem do DataFrame estar disponível e correto
        normalizar_status(df)
        df.to_excel(ARQUIVO_EXCEL_ENTRADA, index=False)
        aplicar_cores_status(ARQUIVO_EXCEL_ENTRADA, df)
        test_api_configuration()
        rodar_agente()
        end_time: datetime.datetime = datetime.datetime.now()
        tempo_total: float = (end_time - start_time).total_seconds()
        logging.info(f"Processo diário do PMO concluído com sucesso! Foram processados {len(df)} projetos em {round(tempo_total, 2)} segundos." )
    except FileNotFoundError:
        logging.error(f" ERRO: Ficheiro CSV de entrada não encontrado: {ARQUIVO_CSV}")
        exit()  # Encerra o script se o arquivo CSV não for encontrado, pois os próximos passos dependem do arquivo Excel estar disponível e formatado corretamente
    except Exception as e:
        logging.error(f"FALHA NO SISTEMA: {str(e)}")
       
    
    

