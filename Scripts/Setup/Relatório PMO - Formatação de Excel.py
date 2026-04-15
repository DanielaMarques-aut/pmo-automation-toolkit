"""PMO Excel Report Formatting and Styling Module.

Automates Excel styling for professional PMO reports. Converts CSV data to
visually formatted Excel with conditional formatting and header styling.

Pipeline:
    1. Load CSV data using pandas
    2. Convert to Excel (.xlsx) format
    3. Apply professional styling with openpyxl
    4. Save formatted output file

Styling Features:
    - Header Row: Dark blue background with white bold text
    - Data Rows: Default formatting (extensible for custom rules)
    - Alignment: Centered headers for readability
    - Output: Professional report suitable for stakeholder distribution

Data Flow:
    CSV → Pandas DataFrame → Excel → OpenPyXL Styling → Formatted Excel

Use Cases:
    - Automated PMO reporting (daily, weekly, monthly)
    - Stakeholder-facing reports that need polish
    - Executive summaries with consistent branding
    - Report templates for distribution

Examples:
    Format single CSV file:
    
    >>> formatar_excel_pmo('relatorio_final.csv')
    >>> # Output: Relatorio_Formatado_PMO.xlsx with dark blue headers
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Optional


def formatar_excel_pmo(file_path: str) -> Optional[str]:
    """Convert CSV to formatted Excel with professional styling.
    
    Two-stage formatting process:
    1. CSV → Excel conversion (pandas)
    2. Style application (openpyxl): headers, colors, fonts
    
    Args:
        file_path (str): Source CSV filename (relative or absolute path).
            Example: "relatorio_final.csv"
            File must exist in current working directory or provided path.
    
    Returns:
        Optional[str]: Output Excel filename if successful.
            Example: "Relatorio_Formatado_PMO.xlsx"
            Returns None if conversion or styling failed.
    
    Raises:
        No exceptions raised to caller. All errors caught and logged:
        - FileNotFoundError: Source CSV doesn't exist
        - ValueError: CSV format invalid or unreadable
        - openpyxl.exceptions: Excel structure errors
    
    Notes:
        TWO-LIBRARY APPROACH:
        - pandas: Efficient CSV → Excel conversion, handles large files
        - openpyxl: Low-level Excel styling (colors, fonts, alignment)
        
        Pandas is easier for bulk operations but doesn't support styling.
        OpenPyXL has powerful styling but requires loading saved file.
        Combining both provides best of each approach.
        
        STYLING CONSTANTS:
        Dark Blue ("000080") is PMO standard for headers.
        White (#FFFFFF) font ensures readability against dark background.
        Bold font adds visual hierarchy and professional appearance.
        Center alignment improves column header readability.
        
        PRODUCTION ENHANCEMENT:
        Current version styles only header row. Extend to support:
        - Alternating row colors for data readability
        - Conditional formatting based on values (KPI thresholds)
        - Column width auto-adjustment based on content
        - Number formatting (currency, percentages, dates)
        - Data validation rules for input fields
    
    Examples:
        Format standard PMO report:
        
        >>> output = formatar_excel_pmo('relatorio_final.csv')
        >>> if output:
        ...     print(f"Report formatted: {output}")
        ... else:
        ...     print("Formatting failed - check CSV file")
        
        Use result in email sending:
        
        >>> excel_file = formatar_excel_pmo('dados.csv')
        >>> if excel_file:
        ...     enviar_alerta_pmo(
        ...         "Weekly Report",
        ...         "Weekly metrics attached.",
        ...         excel_file
        ...     )
    """
    output_file: str = 'Relatorio_Formatado_PMO.xlsx'
    
    try:
        # ================================================================
        # STAGE 1: CSV LOADING AND CONVERSION
        # ================================================================
        print(f"🎨 Formatting visually: {file_path}")
        
        # Load CSV data
        df = pd.read_csv(file_path)
        print(f"✅ Loaded {len(df)} rows from {file_path}")
        
        # Convert to Excel format (creates workbook file)
        df.to_excel(output_file, index=False)
        print(f"✅ Converted to Excel: {output_file}")
        
        # ================================================================
        # STAGE 2: STYLING WITH OPENPYXL
        # ================================================================
        # Open saved Excel file for styling
        wb = load_workbook(output_file)
        ws = wb.active
        
        # Define header styling
        azul_escuro = PatternFill(
            start_color="000080",
            end_color="000080",
            fill_type="solid"
        )
        fonte_branca = Font(
            color="FFFFFF",
            bold=True
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Apply styling to header row (row 1)
        for cell in ws[1]:
            cell.fill = azul_escuro
            cell.font = fonte_branca
            cell.alignment = center_align
        
        print(f"✅ Header styling applied")
        
        # Save styled workbook
        wb.save(output_file)
        print(f"✨ Professional report generated: {output_file}")
        
        return output_file
        
    except FileNotFoundError:
        print(f"❌ Error: Source file '{file_path}' not found!")
        return None
    except ValueError as ve:
        print(f"❌ CSV format error: {str(ve)}")
        return None
    except Exception as e:
        print(f"❌ Error formatting Excel: {type(e).__name__}: {str(e)}")
        return None


if __name__ == "__main__":
    # Format sample report
    print("--- PMO Excel Formatter ---\n")
    result = formatar_excel_pmo('relatorio_final.csv')
    
    if result:
        print(f"\n✅ Output: {result}")
    else:
        print("\n❌ Formatting failed")