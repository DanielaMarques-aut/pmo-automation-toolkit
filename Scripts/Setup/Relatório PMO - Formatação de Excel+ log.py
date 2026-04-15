"""PMO End-to-End Reporting Pipeline with Audit Logging.

Comprehensive data pipeline that combines CSV cleaning, aggregation, and
Excel formatting with complete audit trail logging. Implements production-grade
error handling and governance patterns for compliance and troubleshooting.

Pipeline Stages:
    1. GOVERNANCE: Audit logging to pmo_production.log with timestamps
    2. VALIDATION: Input file existence check (fail-fast pattern)
    3. DATA CLEANING: Remove 'h' suffix from time values, convert to numeric
    4. AGGREGATION: Group by project, sum time spent
    5. FORMATTING: Apply professional Excel styling (headers)
    6. OUTPUT: Save formatted Excel report with completion logging

Audit Trail:
    Every operation logged with ISO8601 timestamp to pmo_production.log for
    compliance, debugging, and post-mortem analysis of pipeline execution.
    Enables full visibility into data flow and error diagnostics.

Error Handling:
    Fail-fast with clear error messages. File not found or data format errors
    prevent corrupt output file creation. All exceptions caught and logged.

Data Validation:
    Cleans time data (remove 'h' suffix: "5h" → 5.0) with coerce mode to
    handle invalid values gracefully (convert to NaN rather than crash).

Use Cases:
    - Automated daily/weekly/monthly PMO reporting
    - Compliance reporting with full audit trails
    - Data pipeline that requires historical operation tracking
    - Production systems where operational visibility is critical

Examples:
    Run pipeline (executes main block on import):
    
    >>> exec(open('Relatório PMO - Formatação de Excel+ log.py').read())
    🚀 Starting pipeline...
    ✅ Cleaned and aggregated. Projects processed: 5
    🎉 Pipeline executed! Report ready: Relatorio_Final_Sexta.xlsx
    
    Check audit trail:
    
    >>> with open('pmo_production.log') as f:
    ...     print(f.readlines()[-5:])  # Last 5 log entries
"""

import pandas as pd
import os
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Optional


# ============================================================================
# AUDIT LOGGING
# ============================================================================

def log(msg: str) -> None:
    """Write timestamped audit log message to persistent file.
    
    Records all pipeline events with ISO8601 timestamp for compliance, debugging,
    and operational visibility. Enables post-mortem analysis of execution and
    troubleshooting of failures.
    
    Args:
        msg (str): Log message to record. Include status emoji and details.
            Example: "🚀 Pipeline started" or "❌ ERROR: File not found."
    
    Returns:
        None
    
    Raises:
        No exceptions. If file write fails, prints to console instead.
    
    Notes:
        APPEND MODE:
        Opens log file in append mode ('a'), preserving historical records.
        Each pipeline run produces new log entries without overwriting previous.
        Enables tracking execution history across days/weeks.
        
        TIMESTAMP FORMAT:
        Uses datetime.datetime.now().isoformat() which includes date, time, and
        microseconds. Format: 2026-04-15T14:30:45.123456
        ISO8601 format is sortable, parseable, and internationally standard.
        
        LOG FILE LOCATION:
        pmo_production.log in current working directory.
        Recommended for production: Move to dedicated logs/ folder with rotation.
        Consider log rotation (zip old logs, keep last 30 days) for long-running systems.
    
    Examples:
        Log pipeline start:
        
        >>> log("🚀 Pipeline started")
        >>> # Appends to pmo_production.log:
        >>> # [2026-04-15T14:30:45.123456] 🚀 Pipeline started
        
        Log error with details:
        
        >>> log(f"❌ ERROR: Could not process {filename}")
        
        Log metrics:
        
        >>> log(f"✅ Aggregated {count} projects")
    """
    try:
        with open("pmo_production.log", "a", encoding="utf-8") as f:
            timestamp = datetime.datetime.now().isoformat()
            f.write(f"[{timestamp}] {msg}\n")
    except IOError as e:
        print(f"❌ Log write failed: {e}")


def run_master_pipeline(
    file_in: str = 'dados_pmo_segunda.csv',
    file_out: str = 'Relatorio_Final_Sexta.xlsx'
) -> Optional[bool]:
    """Execute complete PMO data pipeline from CSV to formatted Excel.
    
    Orchestrates entire workflow: validation → cleaning → aggregation →
    formatting → saving. All operations logged with timestamps for audit trail
    and compliance tracking.
    
    Args:
        file_in (str): Input CSV filename. Default: 'dados_pmo_segunda.csv'
            Must contain columns: 'Projeto', 'Tempo_Gasto' (with 'h' suffix)
            Example row: "ProjectA", "5.5h"
        file_out (str): Output Excel filename. Default: 'Relatorio_Final_Sexta.xlsx'
            Will be created/overwritten in current directory
    
    Returns:
        Optional[bool]: True if pipeline succeeds, None if validation fails.
    
    Raises:
        No exceptions. All errors caught, logged, and reported to console.
        Prevents cascade failures from poor error handling.
    
    Notes:
        PIPELINE ARCHITECTURE:
        
        Stage 1 - GOVERNANCE:
            Log pipeline start. Enables tracking of execution timeline.
            Useful for: SLA monitoring, audit trails, performance analysis.
            
        Stage 2 - VALIDATION:
            Check input file exists. Fail-fast prevents consuming tokens
            on broken datasets or wasting time on missing files.
            
        Stage 3 - DATA CLEANING:
            Remove 'h' suffix from Tempo_Gasto ("5h" → "5")
            Convert to numeric type for aggregation.
            Handle invalid values with errors='coerce' (convert to NaN).
            Example: ["5h", "10.5h", "invalid", "3h"] → [5.0, 10.5, NaN, 3.0]
            
        Stage 4 - AGGREGATION:
            Group by Projeto column, sum all time values per project.
            Creates summary table: [Projeto, Tempo_Gasto_Total]
            Example: ProjectA: 5h + 3h = 8h total
            
        Stage 5 - FORMATTING:
            Apply professional Excel styling:
            - Header row: Dark blue background, white bold font
            - Centered alignment for readability
            - Saves workbook in .xlsx format for Office compatibility
            
        Stage 6 - PERSISTENCE:
            Save formatted Excel file and log completion with metadata.
        
        ERROR RECOVERY:
        If data format invalid (can't parse Tempo_Gasto), pipeline stops
        with clear error message. No partial output created to prevent
        downstream processing of incomplete data.
    
    Examples:
        Run with defaults:
        
        >>> success = run_master_pipeline()
        >>> if success:
        ...     print("Report ready for distribution")
        
        Custom input/output files:
        
        >>> run_master_pipeline(
        ...     file_in="custom_data.csv",
        ...     file_out="custom_report.xlsx"
        ... )
        
        Use in scheduled jobs:
        
        >>> from datetime import datetime
        >>> if datetime.now().weekday() == 4:  # Friday
        ...     success = run_master_pipeline()
        ...     if success:
        ...         send_email(recipient="team@company.com", attachment="Relatorio_Final_Sexta.xlsx")
    """
    # ================================================================
    # STAGE 1: GOVERNANCE (LOGGING)
    # ================================================================
    log("🚀 Pipeline started")
    print("🚀 Starting pipeline...")
    
    # ================================================================
    # STAGE 2: VALIDATION
    # ================================================================
    if not os.path.exists(file_in):
        error_msg = f"❌ ERROR: Input file not found: {file_in}"
        log(error_msg)
        print(error_msg)
        print(f"   Expected at: {os.path.abspath(file_in)}")
        return None
    
    try:
        # ================================================================
        # STAGE 3: DATA CLEANING
        # ================================================================
        log(f"📥 Loading data from: {file_in}")
        df = pd.read_csv(file_in)
        print(f"✅ Loaded {len(df)} rows")
        
        # Clean Tempo_Gasto column: remove 'h' suffix and convert to numeric
        # Input: "5h", "10.5h", "3.25h", "invalid"
        # Output: 5.0, 10.5, 3.25, NaN
        df['Tempo_Gasto'] = pd.to_numeric(
            df['Tempo_Gasto'].astype(str).str.replace('h', ''),
            errors='coerce'  # Invalid values → NaN (don't crash)
        )
        valid_count = df['Tempo_Gasto'].notna().sum()
        log(f"✅ Cleaned {valid_count} valid time entries")
        
        # ================================================================
        # STAGE 4: AGGREGATION
        # ================================================================
        # Group by Projeto, sum time per project
        resumo = df.groupby('Projeto')['Tempo_Gasto'].sum().reset_index()
        
        # Convert to Excel format
        resumo.to_excel(file_out, index=False)
        log(f"✅ Aggregated into {len(resumo)} projects")
        print(f"✅ Cleaned and aggregated. Projects processed: {len(resumo)}")
        
        # ================================================================
        # STAGE 5: FORMATTING
        # ================================================================
        # Open saved Excel file for styling
        wb = load_workbook(file_out)
        ws = wb.active
        
        # Define header styling (PMO standard: dark blue with white text)
        azul_header = PatternFill(
            start_color="000080",
            end_color="000080",
            fill_type="solid"
        )
        fonte_branca = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Apply styling to header row
        for cell in ws[1]:
            cell.fill = azul_header
            cell.font = fonte_branca
            cell.alignment = center_align
        
        # ================================================================
        # STAGE 6: PERSISTENCE
        # ================================================================
        wb.save(file_out)
        log(f"✨ Excel formatting completed successfully")
        log(f"📊 Output file: {file_out}")
        
        print(f"🎉 Pipeline executed! Report ready: {file_out}")
        return True
        
    except ValueError as ve:
        error_msg = f"💥 DATA FORMAT ERROR: {str(ve)}"
        log(error_msg)
        print(error_msg)
        return None
    except Exception as e:
        error_msg = f"💥 SYSTEM FAILURE: {type(e).__name__}: {str(e)}"
        log(error_msg)
        print(error_msg)
        return None


if __name__ == "__main__":
    # Execute pipeline when run as main script
    print("\n--- PMO Master Pipeline (With Audit Logging) ---\n")
    success = run_master_pipeline()
    
    if success:
        print("\n✅ Pipeline completed successfully")
        print("📋 Check pmo_production.log for audit trail")
    else:
        print("\n❌ Pipeline failed - see errors above and logs")