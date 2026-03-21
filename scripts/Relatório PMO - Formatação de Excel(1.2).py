# Relatório PMO - Formatação de Excel
#Learning Python with a fun project: Formatting an Excel report for PMO! 
# This script reads a CSV file, converts it to Excel, and applies some styling to make it look professional.
# The goal is to automate the generation of a visually appealing report that can be easily shared with stakeholders.
# included colors based on status, bold headers, and a clean layout to enhance readability.
#sent email with the formatted report as an attachment to ensure that the PMO team receives the latest insights without any manual effort.
# the email encludes only the new risks detected in the report, making it easier for the team to focus on what matters most.

import json
import os
import smtplib
import logging
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from email.message import EmailMessage
from openpyxl.styles import Font, PatternFill
from dotenv import load_dotenv

# 1. Cores (HEX Codes)
verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
amarelo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
COR_ALVO = vermelho.start_color.rgb[2:]  # Cor vermelha sem o prefixo 'FF' para comparação

 # 2. Estilizar o cabeçalho (Linha 1)
azul_escuro = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
fonte_branca = Font(color="FFFFFF", bold=True)

# 3. CONFIGURAÇÕES 
ARQUIVO_EXCEL_entrada = "Report_Quinta.xlsx"
ARQUIVO_MEMORIA = "alertas_enviados.json"
Arquivo_csv = "dados_pmo_segunda.csv"
ARQUIVO_EXCEL_formatado = 'Relatorio_Formatado_PMO.xlsx'

logging.basicConfig(filename='pmo_report.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Configurações de E-mail 
load_dotenv()
EMAIL_USER = os.getenv("EMAIL_ADDRESS")  # Defina a variável de ambiente EMAIL_USER com seu email
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")  # Defina a variável de ambiente EMAIL_PASSWORD com sua senha
logging.info("Variáveis de ambiente carregadas com sucesso.")
# --- FUNÇÕES DE APOIO ---
def carregar_memoria():
    if os.path.exists(ARQUIVO_MEMORIA):
        with open(ARQUIVO_MEMORIA, 'r') as f: return json.load(f)
    return {}

def salvar_memoria(memoria):
    with open(ARQUIVO_MEMORIA, 'w') as f: json.dump(memoria, f, indent=4)

# --- FUNÇOES PRINCIPAIS ---
def aplicar_cores_status(ARQUIVO_EXCEL_entrada,df):    
    logging.info(f"Aplicar cores com base no status: {ARQUIVO_EXCEL_entrada}")
    wb = load_workbook(ARQUIVO_EXCEL_entrada)
    ws = wb.active
    #Aplicar cores com base no status da tarefa
    #Aplicar vermelho para "Atrasado"
    df_riscos = df[df["Status"] == "Atrasado"]
    for index in df_riscos.index:
        ws.cell(row=index + 2, column=3).fill = vermelho  # +2 para pular o cabeçalho e ajustar o índice
        #Aplicar cor verde para os itens "Concluido"
    df_concluidos = df[df["Status"] == "Concluído"]
    for index in df_concluidos.index:
        ws.cell(row=index + 2, column=3).fill = verde  # +2 para pular o cabeçalho e ajustar o índice
    df_outros = df[~df["Status"].isin(["Atrasado", "Concluído"])]
    for index in df_outros.index:
        ws.cell(row=index + 2, column=3).fill = amarelo  # +2 para pular o cabeçalho e ajustar o índice
        logging.info(f"A formatar visualmente o cabeçalho do relatório: {ARQUIVO_EXCEL_entrada }")
    for cell in ws[1]: # Para cada célula na primeira linha
        cell.fill = azul_escuro
        cell.font = fonte_branca
    wb.save(ARQUIVO_EXCEL_formatado)
    logging.info(f"Cores aplicadas com base no status. Relatório atualizado: {ARQUIVO_EXCEL_entrada}")
    logging.info(f"✨ Report Profissional gerado: {ARQUIVO_EXCEL_formatado}")

if EMAIL_USER is None or EMAIL_PASSWORD is None:
    logging.error(f"ERRO CRÍTICO: O Python não encontrou o arquivo .env ou as variáveis!")
    logging.debug(f"DEBUG: EMAIL_ADDRESS encontrado? {'Sim' if EMAIL_USER else 'Não'}")
    logging.debug(f"DEBUG: EMAIL_PASSWORD encontrado? {'Sim' if EMAIL_PASSWORD else 'Não'}")
    # Encerra o script para não dar erro de 'NoneType'
    exit()
# --- LOGICA PRINCIPAL DO AGENTE ---
def rodar_agente():
    memoria = carregar_memoria()
    wb = openpyxl.load_workbook(ARQUIVO_EXCEL_formatado, data_only=True)
    ws = wb.active
    novos_alertas = []

    for row in range(2, ws.max_row + 1):
        nome = ws.cell(row=row, column=1).value
        cor = ws.cell(row=row, column=3).fill.start_color.rgb[2:] # ter o RGB sem o prefixo 'FF'
        print (f"Verificando tarefa: {nome} - Cor: {cor}")
        if cor == COR_ALVO and nome not in memoria:
            print(f"Novo risco detectado: {nome}")
            novos_alertas.append({"tarefa": nome})
            memoria[nome] = "NOTIFICADO"

    if novos_alertas:
        salvar_memoria(memoria)
        logging.warning(f"{len(novos_alertas)} novos riscos encontrados. A guardar... ")
        enviar_email(novos_alertas)
    else:
        # Feedback caso o dia esteja tranquilo
        print("✅ Tudo em dia! Nenhum novo risco detectado.")
        logging.info("Nenhum novo risco encontrado para notificação.")
    logging.info("Agente de Riscos PMO executado com sucesso!")
def enviar_email(alertas):
    msg = EmailMessage()
    msg['Subject'] = f"Relatório de Riscos - {len(alertas)} itens"
    msg['From'] = EMAIL_USER
    msg['To'] = EMAIL_USER
    conteudo = "Olá,\n\nSegue o resumo dos riscos detectados no relatório PMO.\n\n Atenciosamente,\n Agente de Riscos PMO,\n\nResumo de Riscos\n\n"   
    msg.set_content(conteudo)
    for a in alertas:
        conteudo += f"Tarefa: {a['tarefa']}\n"
    msg.set_content(conteudo)
    msg.add_attachment(open(ARQUIVO_EXCEL_formatado, 'rb').read(), maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=ARQUIVO_EXCEL_formatado)
    
    
    with smtplib.SMTP("smtp.gmail.com", 587) as s:
        try:
            s.starttls()
            s.login(EMAIL_USER, EMAIL_PASSWORD)
            s.send_message(msg)
            logging.info("📧 Email de alerta enviado com sucesso!")
        except Exception as e:
            logging.error(f"💥 FALHA AO ENVIAR EMAIL: {str(e)}" )
if __name__ == "__main__":
    # Aqui colocamos a ordem dos passos que o script deve seguir
    logging.info("🚀 Iniciando processo diário do PMO...")
    try:
        df = pd.read_csv(Arquivo_csv)
        df.to_excel(ARQUIVO_EXCEL_entrada, index=False)
        aplicar_cores_status(ARQUIVO_EXCEL_entrada, df)
        rodar_agente()
        logging.info("Processo diário do PMO concluído com sucesso!")
    except FileNotFoundError:
        logging.error(f"❌ ERRO: Ficheiro CSV de entrada não encontrado: {Arquivo_csv}")
        exit()  # Encerra o script se o arquivo CSV não for encontrado, pois os próximos passos dependem do arquivo Excel estar disponível e formatado corretamente
    except Exception as e:
        logging.error(f"💥 FALHA NO SISTEMA: {str(e)}")
